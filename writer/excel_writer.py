"""Write matched prices back into the LV Excel file."""
import shutil
from pathlib import Path
from typing import List, Optional
import openpyxl
from openpyxl.styles import PatternFill, Font
import config
from matching.price_validator import validate_match


# Fills
FILL_GREEN = PatternFill(start_color=config.COLOR_GREEN, end_color=config.COLOR_GREEN, fill_type="solid")
FILL_ORANGE = PatternFill(start_color=config.COLOR_ORANGE, end_color=config.COLOR_ORANGE, fill_type="solid")
FILL_RED = PatternFill(start_color=config.COLOR_RED, end_color=config.COLOR_RED, fill_type="solid")
SUPPLIER_FONT = Font(size=8, color="666666")

PRIORITY = {config.COLOR_GREEN: 1, config.COLOR_ORANGE: 2, config.COLOR_RED: 3}


def write_prices_to_lv(
    lv_path: str,
    matches: list,
    output_path: Optional[str] = None,
    sheet_name: str = "Kalkulation",
    source_type: str = "angebot",  # "angebot" | "pdb" | "internet"
) -> dict:
    """Write matched prices into the LV Excel.
    
    NEW LOGIC:
    - NU (Nachunternehmer) → Column M (EP EK)
    - Stoffe (Materials) → Components to F7-F1 → Sum to Column W (Stoffe-Kosten)
    
    Args:
        lv_path: Path to the LV Excel file
        matches: List of resolved matches [{
            row: int,              # Excel row number
            column: "X" or "M",   # Stoffe or NU (X=Stoffe, M=NU)
            ep: float,             # Unit price
            supplier: str,         # Supplier name
            oz: str,               # Position number (for report)
            bezeichnung: str,      # Description (for report)
            explanation: str,      # Price calculation explanation
            warning: str,          # Optional warning
            components: [{         # For multi-material positions (optional)
                name: str,         # Component name (e.g., "L-Stein", "Unterbeton")
                ep: float,         # Component unit price
                einheit: str,      # Component unit
            }]
        }]
        output_path: Where to save (default: adds _filled suffix)
        sheet_name: Sheet to modify
        source_type: Determines color (angebot=green, pdb=orange, internet=red)
    
    Returns: Report dict with stats
    """
    lv_path = Path(lv_path)
    if not output_path:
        output_path = lv_path.parent / f"{lv_path.stem}_filled{lv_path.suffix}"
    output_path = Path(output_path)
    
    # ALWAYS create backup first
    backup_path = lv_path.parent / f"{lv_path.stem}_backup{lv_path.suffix}"
    if not backup_path.exists():
        shutil.copy2(lv_path, backup_path)
    
    # Copy to output path for editing
    shutil.copy2(lv_path, output_path)
    
    # Select fill color based on source
    fill = {"angebot": FILL_GREEN, "pdb": FILL_ORANGE, "internet": FILL_RED}[source_type]
    color_hex = {"angebot": config.COLOR_GREEN, "pdb": config.COLOR_ORANGE, "internet": config.COLOR_RED}[source_type]
    color_priority = PRIORITY[color_hex]
    
    wb = openpyxl.load_workbook(str(output_path))
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    
    report = {
        "written": [],
        "skipped_priority": [],
        "skipped_formula": [],
        "skipped_cheaper": [],
        "replaced": [],
        "warnings": [],
        "stats": {"total_matches": len(matches), "written": 0, "skipped": 0, "replaced": 0},
    }
    
    col = config.LV_COLUMNS
    
    for match in matches:
        # Validate price before writing
        is_valid, reason = validate_match(match)
        if not is_valid:
            report["warnings"].append({"oz": match.get("oz", ""), "warning": f"Nicht geschrieben: {reason}"})
            report["stats"]["skipped"] += 1
            continue

        row = match["row"]
        is_nu = match.get("column") == "M"
        ep = match["ep"]

        # Safety net: force NU for positions that are clearly services/work
        if not is_nu:
            _NU_KEYWORDS = ("dichtheitsprüfung", "dichtheitspr", "inspektion",
                            "tv-befahrung", "kamerabefahrung", "kontrollprüfung",
                            "druckprüfung", "dokumentation", "vermessung",
                            "gutachten", "abnahme", "protokoll",
                            "baustelleneinrichtung", "verkehrssicherung",
                            "absperrung", "beschilderung")
            bez = match.get("bezeichnung", "").lower()
            if any(kw in bez for kw in _NU_KEYWORDS):
                is_nu = True
                match["column"] = "M"
        supplier = match.get("supplier", "")
        components = match.get("components", [])  # Multi-material components
        
        if ep is None or ep <= 0:
            continue
        
        if is_nu:
            # NU (Nachunternehmer) → Column M (EP EK)
            _write_single_price(
                ws, row, col["nu_ep"], col["lieferant"],
                ep, supplier, fill, color_priority, match, report
            )
        else:
            # Stoffe (Materials) → F-columns + Stoffe-Kosten
            if components and len(components) > 1:
                # Multi-material: write components to F7-F1, sum to W
                _write_multi_material(
                    ws, row, col, components, supplier, fill, match, report
                )
            else:
                # Single material: write to F7, copy to Stoffe-Kosten
                _write_single_material(
                    ws, row, col, ep, supplier, fill, match, report
                )
        
        if match.get("warning"):
            report["warnings"].append({
                "oz": match.get("oz", ""),
                "warning": match.get("warning"),
            })
    
    # Save
    wb.save(str(output_path))
    wb.close()
    
    report["output_file"] = str(output_path)
    report["backup_file"] = str(backup_path)
    
    return report


def _write_single_price(ws, row, target_col, lieferant_col, ep, supplier, fill, color_priority, match, report):
    """Write a single price to a cell (for NU)."""
    cell = ws.cell(row=row, column=target_col)
    supplier_cell = ws.cell(row=row, column=lieferant_col)
    
    # Check if cell has a formula (NEVER overwrite formulas)
    if isinstance(cell.value, str) and str(cell.value).startswith("="):
        report["skipped_formula"].append({
            "oz": match.get("oz", ""),
            "reason": f"Formel in Zelle ({cell.value[:30]}...)",
        })
        report["stats"]["skipped"] += 1
        return
    
    # Check existing color priority
    existing_color = _get_fill_hex(cell.fill)
    existing_priority = PRIORITY.get(existing_color, 99)
    
    if existing_priority < color_priority and cell.value:
        existing_value = _to_float(cell.value)
        if existing_value and existing_value > 0:
            if ep < existing_value:
                # New price is cheaper - replace!
                old_val = existing_value
                cell.value = ep
                cell.fill = fill
                supplier_cell.value = supplier
                supplier_cell.font = SUPPLIER_FONT
                
                savings = round(old_val - ep, 2)
                report["replaced"].append({
                    "oz": match.get("oz", ""),
                    "bezeichnung": match.get("bezeichnung", ""),
                    "old_price": old_val,
                    "new_price": ep,
                    "savings": savings,
                    "supplier": supplier,
                })
                report["stats"]["replaced"] += 1
                return
            else:
                report["skipped_cheaper"].append({
                    "oz": match.get("oz", ""),
                    "existing": existing_value,
                    "new": ep,
                    "reason": f"Bestehender Preis {existing_value:.2f}€ günstiger als {ep:.2f}€",
                })
                report["stats"]["skipped"] += 1
                return
        
        report["skipped_priority"].append({
            "oz": match.get("oz", ""),
            "reason": f"Höhere Priorität (bestehend: {existing_color})",
        })
        report["stats"]["skipped"] += 1
        return
    
    # Write the price
    cell.value = ep
    cell.fill = fill
    supplier_cell.value = supplier
    supplier_cell.font = SUPPLIER_FONT
    
    report["written"].append({
        "oz": match.get("oz", ""),
        "bezeichnung": match.get("bezeichnung", ""),
        "ep": ep,
        "column": match.get("column", "M"),
        "supplier": supplier,
        "explanation": match.get("explanation", ""),
    })
    report["stats"]["written"] += 1


def _write_multi_material(ws, row, col, components, supplier, fill, match, report):
    """Write multi-material components to F7-F2 and sum to Stoffe-Kosten."""
    # F-column mapping: f7, f6, f5, f4, f3, f2 (NOT f1 - leave empty)
    f_cols = [col.get("f7"), col.get("f6"), col.get("f5"), col.get("f4"), 
              col.get("f3"), col.get("f2")]
    
    total_ep = 0
    written_components = []
    
    # Write each component to F-columns (up to 6 components, F1 is left empty)
    for i, comp in enumerate(components[:6]):
        comp_ep = comp.get("ep", 0)
        comp_name = comp.get("name", f"Komponente {i+1}")
        comp_unit = comp.get("einheit", "")
        
        if comp_ep > 0 and f_cols[i]:
            cell = ws.cell(row=row, column=f_cols[i])
            cell.value = comp_ep
            cell.fill = fill
            total_ep += comp_ep
            written_components.append(f"{comp_name}: {comp_ep:.2f}€/{comp_unit}")
    
    # Write total to Stoffe-Kosten (column W)
    if total_ep > 0 and col.get("stoffe_kosten"):
        total_cell = ws.cell(row=row, column=col["stoffe_kosten"])
        total_cell.value = round(total_ep, 2)
        total_cell.fill = fill
        
        # Write supplier to lieferant column
        supplier_cell = ws.cell(row=row, column=col.get("lieferant", 20))
        supplier_cell.value = supplier
        supplier_cell.font = SUPPLIER_FONT
    
    report["written"].append({
        "oz": match.get("oz", ""),
        "bezeichnung": match.get("bezeichnung", ""),
        "ep": round(total_ep, 2),
        "column": "X",
        "supplier": supplier,
        "explanation": match.get("explanation", "") + " | " + "; ".join(written_components),
        "components": len(written_components),
    })
    report["stats"]["written"] += 1


def _write_single_material(ws, row, col, ep, supplier, fill, match, report):
    """Write single material to F7 and Stoffe-Kosten."""
    # Write to F7 (first component column)
    if col.get("f7"):
        f7_cell = ws.cell(row=row, column=col["f7"])
        f7_cell.value = ep
        f7_cell.fill = fill
    
    # Write to Stoffe-Kosten (column W)
    if col.get("stoffe_kosten"):
        total_cell = ws.cell(row=row, column=col["stoffe_kosten"])
        total_cell.value = ep
        total_cell.fill = fill
    
    # Write supplier
    supplier_cell = ws.cell(row=row, column=col.get("lieferant", 20))
    supplier_cell.value = supplier
    supplier_cell.font = SUPPLIER_FONT
    
    report["written"].append({
        "oz": match.get("oz", ""),
        "bezeichnung": match.get("bezeichnung", ""),
        "ep": ep,
        "column": "X",
        "supplier": supplier,
        "explanation": match.get("explanation", ""),
    })
    report["stats"]["written"] += 1


def _get_fill_hex(fill) -> Optional[str]:
    try:
        if fill and fill.fgColor and fill.fgColor.rgb:
            rgb = str(fill.fgColor.rgb)
            if rgb != "00000000" and len(rgb) >= 6:
                return rgb[-6:].upper()
    except Exception:
        pass
    return None


def _to_float(val) -> Optional[float]:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        if isinstance(val, str):
            cleaned = val.replace(".", "").replace(",", ".").replace("€", "").strip()
            try:
                return float(cleaned)
            except ValueError:
                return None
        return None
