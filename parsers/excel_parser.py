"""Parse LV Excel files and offer spreadsheets."""
from pathlib import Path
from typing import Optional
import openpyxl
from openpyxl.utils import get_column_letter
import config


def parse_lv_excel(filepath: str, sheet_name: str = "Kalkulation") -> dict:
    """Read an LV Excel and return structured position data.
    
    Returns: {
        positions: [{row, oz, bezeichnung, menge, einheit, stoffe_ep, nu_ep, lieferant, has_formula_x, has_formula_m}],
        stats: {total, filled_x, filled_m, empty},
        columns: {detected column mapping}
    }
    """
    filepath = Path(filepath)
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    
    if sheet_name not in wb.sheetnames:
        # Try to find the right sheet
        for name in wb.sheetnames:
            if "kalk" in name.lower() or "lv" in name.lower():
                sheet_name = name
                break
        else:
            sheet_name = wb.sheetnames[0]
    
    ws = wb[sheet_name]
    
    # Also load with formulas to check for formula cells
    wb_formulas = openpyxl.load_workbook(str(filepath))
    ws_formulas = wb_formulas[sheet_name]
    
    col = config.LV_COLUMNS
    positions = []
    
    for row_idx in range(1, ws.max_row + 1):
        oz = ws.cell(row=row_idx, column=col["oz"]).value
        bez = ws.cell(row=row_idx, column=col["bezeichnung"]).value
        
        if not oz or not bez:
            continue
        
        # Skip header rows
        oz_str = str(oz).strip()
        if oz_str.lower() in ("oz", "pos", "pos.", "position", "nr", "nr."):
            continue
        
        menge = ws.cell(row=row_idx, column=col["menge"]).value
        einheit = ws.cell(row=row_idx, column=col["einheit"]).value

        # Skip group headers / Titel rows (no menge AND no einheit = not a real position)
        if not menge and not einheit:
            continue

        stoffe_ep = ws.cell(row=row_idx, column=col["stoffe_kosten"]).value
        nu_ep = ws.cell(row=row_idx, column=col["nu_ep"]).value
        lieferant = ws.cell(row=row_idx, column=col["lieferant"]).value

        # Check if cells contain formulas
        cell_x = ws_formulas.cell(row=row_idx, column=col["stoffe_kosten"])
        cell_m = ws_formulas.cell(row=row_idx, column=col["nu_ep"])
        has_formula_x = isinstance(cell_x.value, str) and cell_x.value.startswith("=")
        has_formula_m = isinstance(cell_m.value, str) and cell_m.value.startswith("=")
        
        # Check existing fill color
        fill_x = ws.cell(row=row_idx, column=col["stoffe_kosten"]).fill
        fill_m = ws.cell(row=row_idx, column=col["nu_ep"]).fill
        color_x = _get_fill_hex(fill_x)
        color_m = _get_fill_hex(fill_m)
        
        try:
            menge_float = float(menge) if menge else 0
        except (ValueError, TypeError):
            menge_float = 0
        
        positions.append({
            "row": row_idx,
            "oz": oz_str,
            "bezeichnung": str(bez).strip(),
            "menge": menge_float,
            "einheit": str(einheit).strip() if einheit else "",
            "stoffe_ep": _to_float(stoffe_ep),
            "nu_ep": _to_float(nu_ep),
            "lieferant": str(lieferant).strip() if lieferant else "",
            "has_formula_x": has_formula_x,
            "has_formula_m": has_formula_m,
            "color_x": color_x,
            "color_m": color_m,
        })
    
    wb.close()
    wb_formulas.close()
    
    filled_x = sum(1 for p in positions if p["stoffe_ep"] and p["stoffe_ep"] > 0)
    filled_m = sum(1 for p in positions if p["nu_ep"] and p["nu_ep"] > 0)
    empty = sum(1 for p in positions if not p["stoffe_ep"] and not p["nu_ep"])
    
    return {
        "filename": filepath.name,
        "sheet": sheet_name,
        "positions": positions,
        "stats": {
            "total": len(positions),
            "filled_stoffe": filled_x,
            "filled_nu": filled_m,
            "empty": empty,
        }
    }


def parse_offer_excel(filepath: str) -> dict:
    """Parse an Excel offer/Angebot. Auto-detect structure."""
    filepath = Path(filepath)
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    ws = wb.active
    
    # Try to detect columns by header row
    headers = {}
    for row_idx in range(1, min(10, ws.max_row + 1)):
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                val_lower = str(val).lower().strip()
                if any(k in val_lower for k in ("pos", "nr", "artikel")):
                    headers["pos"] = col_idx
                elif any(k in val_lower for k in ("bezeichnung", "beschreibung", "text", "artikel")):
                    headers["text"] = col_idx
                elif any(k in val_lower for k in ("menge", "anzahl")):
                    headers["menge"] = col_idx
                elif any(k in val_lower for k in ("einheit", "me", "eh")):
                    headers["einheit"] = col_idx
                elif any(k in val_lower for k in ("ep", "einzelpreis", "e-preis", "preis/einheit")):
                    headers["ep"] = col_idx
                elif any(k in val_lower for k in ("gp", "gesamtpreis", "gesamt", "betrag")):
                    headers["gp"] = col_idx
        if len(headers) >= 3:
            break
    
    items = []
    start_row = max(r for r in range(1, 10) if any(
        ws.cell(row=r, column=c).value for c in headers.values()
    )) + 1 if headers else 2
    
    for row_idx in range(start_row, ws.max_row + 1):
        text_val = ws.cell(row=row_idx, column=headers.get("text", 2)).value
        if not text_val:
            continue
        
        pos = ws.cell(row=row_idx, column=headers.get("pos", 1)).value
        menge = ws.cell(row=row_idx, column=headers.get("menge", 3)).value
        einheit = ws.cell(row=row_idx, column=headers.get("einheit", 4)).value
        ep = ws.cell(row=row_idx, column=headers.get("ep", 5)).value
        gp = ws.cell(row=row_idx, column=headers.get("gp", 6)).value
        
        items.append({
            "pos": str(pos).strip() if pos else "",
            "text": str(text_val).strip(),
            "menge": _to_float(menge),
            "einheit": str(einheit).strip() if einheit else "",
            "ep": _to_float(ep),
            "gp": _to_float(gp),
        })
    
    wb.close()
    
    return {
        "filename": filepath.name,
        "headers_detected": headers,
        "items": items,
        "total_items": len(items),
    }


def _to_float(val) -> Optional[float]:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        # Try German number format
        if isinstance(val, str):
            cleaned = val.replace(".", "").replace(",", ".").replace("€", "").replace(" ", "")
            try:
                return float(cleaned)
            except ValueError:
                return None
        return None


def _get_fill_hex(fill) -> Optional[str]:
    """Extract hex color from cell fill."""
    try:
        if fill and fill.fgColor and fill.fgColor.rgb:
            rgb = str(fill.fgColor.rgb)
            if rgb != "00000000" and len(rgb) >= 6:
                return rgb[-6:]  # Last 6 chars
    except Exception:
        pass
    return None
